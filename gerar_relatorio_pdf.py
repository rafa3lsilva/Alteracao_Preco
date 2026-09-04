import os
import re
import sys
import json
import warnings
import argparse

# Suprime avisos do openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import openpyxl
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.pdfgen import canvas

ARQUIVO_HISTORICO = ".historico_alteracoes.json"

# Seções e palavras-chave de Hortifrúti que devem ser alertadas/ignoradas
PALAVRAS_CHAVE_HORTIFRUTI = ['HORTIFRUTI', 'HORTIFRUT', 'HORTFRUT', 'FLV', 'HORTI FRUTI', 'HORTI-FRUTI', 'HORTIFRUTICOLA']

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(12 * mm, 11 * mm, 198 * mm, 11 * mm)
        
        agora = datetime.now().strftime("%d/%m/%Y às %H:%M")
        self.drawString(12 * mm, 7 * mm, f"Relatório de Alteração de Preços • Gerado em {agora}")
        
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(198 * mm, 7 * mm, page_text)
        self.restoreState()


def limpar_texto(texto):
    if not texto:
        return ""
    txt = str(texto).strip()
    txt = txt.replace("\ufffd", "").replace("", "")
    return txt


def verificar_conformidade_planilha(caminho_xlsx):
    """
    Verifica se a planilha segue os padrões recomendados:
    1. Deve ser do tipo ANALÍTICO (não Sintético/Simplificado)
    2. Deve estar SEM o setor Hortifrúti
    """
    if isinstance(caminho_xlsx, str) and not os.path.exists(caminho_xlsx):
        return []

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    alertas = []
    
    # 1. Checagem de tipo (Analítico vs Sintético/Simplificado)
    sheet_name = ws.title.lower()
    titulo_relatorio = str(rows[0][0]) if rows and rows[0][0] else ""
    
    is_analitico = "ANALITICO" in titulo_relatorio.upper() or "analitico" in sheet_name
    if not is_analitico:
        alertas.append(
            "[!] TIPO DE RELATÓRIO:\n"
            "    A planilha exportada está no formato SIMPLIFICADO / SINTÉTICO.\n"
            "    * O formato correto no Varejofácil deve ser: LISTAGEM DE PREÇOS - ANALÍTICO\n"
            "      (para conter o histórico de horários das alterações)."
        )

    # 2. Checagem de Hortifrúti
    tem_hortifruti = False
    for r in rows:
        for c in r:
            if c:
                c_upper = str(c).upper()
                if any(kw in c_upper for kw in PALAVRAS_CHAVE_HORTIFRUTI):
                    tem_hortifruti = True
                    break
        if tem_hortifruti:
            break

    if tem_hortifruti:
        alertas.append(
            "[!] SETOR HORTIFRÚTI DETECTADO:\n"
            "    O setor de HORTIFRÚTI foi incluído na exportação da planilha.\n"
            "    * A orientação da loja é gerar o relatório SEM o setor de Hortifrúti."
        )

    return alertas


def extrair_dados_xlsx(caminho_xlsx, ignorar_secoes=None):
    if isinstance(caminho_xlsx, str) and not os.path.exists(caminho_xlsx):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_xlsx}")

    if ignorar_secoes is None:
        ignorar_secoes = PALAVRAS_CHAVE_HORTIFRUTI

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    empresa = "SUPERMERCADO JEAN LTDA"
    data_relatorio = datetime.now().strftime("%d/%m/%Y")
    
    for r in rows[:5]:
        for c in r:
            if c is not None:
                c_str = str(c)
                if "SUPERMERCADO" in c_str.upper():
                    empresa = limpar_texto(c_str)
                if "Período:" in c_str or "Perodo:" in c_str or "Loja:" in c_str:
                    dates = re.findall(r'\d{2}/\d{2}/\d{4}', c_str)
                    if len(dates) >= 2:
                        if dates[0] == dates[1]:
                            data_relatorio = dates[0]
                        else:
                            data_relatorio = f"{dates[0]} a {dates[1]}"
                    elif len(dates) == 1:
                        data_relatorio = dates[0]

    current_section = "GERAL"
    secao_counter = 0
    data_by_section = {}

    for r_idx, r in enumerate(rows):
        # 1. Verifica se a linha define uma Seção
        sec_found = None
        for c in r:
            if c is not None:
                c_str = str(c).strip()
                if "Seção:" in c_str or "Seo:" in c_str or c_str.startswith("Se"):
                    non_none = [str(x).strip() for x in r if x is not None]
                    name_parts = []
                    for part in non_none:
                        if part not in ["Seção:", "Seo:"] and not part.startswith("Se"):
                            name_parts.append(part)
                    if name_parts:
                        sec_found = " ".join(name_parts)
                    else:
                        secao_counter += 1
                        sec_found = f"SEÇÃO {secao_counter}"
                    break
        
        if sec_found:
            sec_found = limpar_texto(sec_found)
            # Ignora se for Hortifruti
            deve_ignorar = any(ign.upper() in sec_found.upper() for ign in ignorar_secoes)
            if deve_ignorar:
                current_section = None
                continue
            
            current_section = sec_found
            if current_section not in data_by_section:
                data_by_section[current_section] = []
            continue

        if current_section is None:
            continue

        # 2. Verifica se a linha define um Produto
        for c in r:
            if c is not None:
                c_str = str(c).strip()
                m = re.match(r'^\s*(\d{3,14})\s*-\s*(.+)$', c_str)
                if m:
                    prod_code = m.group(1).strip()
                    prod_name = limpar_texto(m.group(2))
                    clean_code = str(int(prod_code)) if prod_code.isdigit() else prod_code
                    
                    non_none = [x for x in r if x is not None]
                    
                    money_vals = [str(x).strip() for x in non_none if "R$" in str(x)]
                    time_vals = [str(x).strip() for x in non_none if re.match(r'^\d{2}:\d{2}$', str(x).strip())]
                    
                    if len(money_vals) >= 3:
                        venda_str = money_vals[2] # Formato Analítico
                    elif len(money_vals) >= 1:
                        venda_str = money_vals[0] # Formato Simplificado
                    else:
                        venda_str = "R$ 0,00"
                    
                    hora_str = time_vals[0] if time_vals else "--"
                    
                    sec_key = current_section if current_section else "GERAL"
                    if sec_key not in data_by_section:
                        data_by_section[sec_key] = []
                    
                    item_key = f"{clean_code}#{hora_str}#{venda_str}"
                    
                    data_by_section[sec_key].append({
                        'key': item_key,
                        'cod': clean_code,
                        'cod_raw': prod_code,
                        'descricao': prod_name,
                        'venda_atual': venda_str,
                        'hora': hora_str
                    })
                    break

    return {
        'empresa': empresa,
        'periodo': data_relatorio,
        'secoes': data_by_section
    }


try:
    from storage_manager import carregar_historico, salvar_historico, resetar_historico
except ImportError:
    def carregar_historico():
        if os.path.exists(ARQUIVO_HISTORICO):
            try:
                with open(ARQUIVO_HISTORICO, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def salvar_historico(data_str, itens_processados, info_execucao=""):
        historico = carregar_historico()
        if data_str not in historico:
            historico[data_str] = {'itens': [], 'execucoes': []}
        if isinstance(historico[data_str], list):
            historico[data_str] = {'itens': historico[data_str], 'execucoes': []}
        if info_execucao == "primeiro":
            historico[data_str]['itens'] = [it['key'] for it in itens_processados]
        else:
            conjunto = set(historico[data_str]['itens'])
            for it in itens_processados:
                conjunto.add(it['key'])
            historico[data_str]['itens'] = list(conjunto)
        historico[data_str]['execucoes'].append({
            'horario': datetime.now().strftime("%H:%M:%S"),
            'qtd_itens': len(itens_processados),
            'tipo': info_execucao
        })
        try:
            with open(ARQUIVO_HISTORICO, 'w', encoding='utf-8') as f:
                json.dump(historico, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Aviso ao salvar histórico: {e}")

    def resetar_historico():
        if os.path.exists(ARQUIVO_HISTORICO):
            try:
                os.remove(ARQUIVO_HISTORICO)
                print("Histórico de alterações resetado com sucesso!")
            except Exception as e:
                print(f"Erro ao remover arquivo de histórico: {e}")
        else:
            print("Nenhum histórico anterior encontrado para resetar.")


def filtrar_dados(dados, modo="todos", hora_corte=None, apenas_novos=False):
    data_hoje = dados['periodo'] if dados['periodo'] else datetime.now().strftime("%d/%m/%Y")
    if " a " in data_hoje:
        data_hoje = data_hoje.split(" a ")[-1].strip()

    historico_dia = carregar_historico().get(data_hoje, {})
    if isinstance(historico_dia, dict):
        itens_ja_impressos = set(historico_dia.get('itens', []))
    else:
        itens_ja_impressos = set(historico_dia)

    secoes_filtradas = {}
    todos_itens_filtrados = []

    for sec_nome, itens in dados['secoes'].items():
        itens_validos = []
        for it in itens:
            h = it['hora']
            k = it['key']
            
            if modo == "segundo" or apenas_novos:
                if k in itens_ja_impressos:
                    continue
            
            if modo == "apos" and hora_corte and h and h != "--" and h < hora_corte:
                continue
            if modo == "ate" and hora_corte and h and h != "--" and h > hora_corte:
                continue

            itens_validos.append(it)
            todos_itens_filtrados.append(it)

        if itens_validos:
            secoes_filtradas[sec_nome] = itens_validos

    return secoes_filtradas, todos_itens_filtrados


def gerar_pdf(dados, caminho_pdf_saida, secoes_filtradas, info_turno="", forcar_quebra_secao=False):
    total_itens = sum(len(itens) for itens in secoes_filtradas.values())
    if total_itens == 0:
        print("\n=======================================================")
        print("  AVISO: Nenhuma alteração encontrada para gerar.")
        print("  Verifique se a planilha contém itens alterados.")
        print("=======================================================\n")
        return False

    doc = SimpleDocTemplate(
        caminho_pdf_saida,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=14 * mm
    )

    styles = getSampleStyleSheet()
    
    style_header_title = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        textColor=colors.HexColor("#0F172A")
    )
    
    style_header_subtitle = ParagraphStyle(
        'HeaderSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#475569")
    )
    
    style_section_title = ParagraphStyle(
        'SectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9.5,
        leading=12,
        textColor=colors.HexColor("#1E293B")
    )
    
    style_cell_code = ParagraphStyle(
        'CellCode',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9,
        alignment=1,
        textColor=colors.HexColor("#334155")
    )
    
    style_cell_desc = ParagraphStyle(
        'CellDesc',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#0F172A")
    )
    
    style_cell_price = ParagraphStyle(
        'CellPrice',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=9.5,
        alignment=2,
        textColor=colors.HexColor("#0F172A")
    )
    
    style_cell_hora = ParagraphStyle(
        'CellHora',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.5,
        leading=9,
        alignment=1,
        textColor=colors.HexColor("#475569")
    )

    style_th = ParagraphStyle(
        'ThStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=9.5,
        textColor=colors.white
    )
    style_th_center = ParagraphStyle('ThStyleCenter', parent=style_th, alignment=1)
    style_th_right = ParagraphStyle('ThStyleRight', parent=style_th, alignment=2)

    story = []
    
    header_table_data = [
        [
            Paragraph(f"<b>{dados['empresa']}</b>", style_header_title),
            Paragraph(f"<b>DATA:</b> {dados['periodo']}", ParagraphStyle('RightSub', parent=style_header_subtitle, alignment=2))
        ],
        [
            Paragraph(f"RELATÓRIO DE ALTERAÇÃO DE PREÇOS{info_turno}", style_header_subtitle),
            Paragraph(f"<b>TOTAL DE PRODUTOS:</b> {total_itens} item(ns)", ParagraphStyle('RightSub2', parent=style_header_subtitle, alignment=2))
        ]
    ]
    
    t_header = Table(header_table_data, colWidths=[115 * mm, 71 * mm])
    t_header.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('TOPPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(t_header)
    story.append(Spacer(1, 2.5 * mm))
    
    col_widths = [22 * mm, 114 * mm, 28 * mm, 22 * mm]
    secoes_lista = list(secoes_filtradas.items())

    for idx, (nome_secao, itens) in enumerate(secoes_lista):
        sec_banner = [
            [
                Paragraph(f"SEÇÃO: <b>{nome_secao.upper()}</b>", style_section_title),
                Paragraph(f"<b>{len(itens)}</b> item(ns)", ParagraphStyle('SecCount', parent=style_section_title, alignment=2, textColor=colors.HexColor("#475569"), fontSize=8))
            ]
        ]
        t_sec = Table(sec_banner, colWidths=[140 * mm, 46 * mm])
        t_sec.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#E2E8F0")),
            ('TOPPADDING', (0,0), (-1,-1), 2.5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
            ('LINELEFT', (0,0), (0,0), 3.5, colors.HexColor("#1E3A8A")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))

        table_data = [
            [
                Paragraph("CÓD.", style_th_center),
                Paragraph("DESCRIÇÃO", style_th),
                Paragraph("VENDA ATUAL", style_th_right),
                Paragraph("HORA", style_th_center)
            ]
        ]

        for item in itens:
            table_data.append([
                Paragraph(item['cod'], style_cell_code),
                Paragraph(item['descricao'], style_cell_desc),
                Paragraph(item['venda_atual'], style_cell_price),
                Paragraph(item['hora'], style_cell_hora)
            ])

        t_itens = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 2.5),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2.5),
            ('TOPPADDING', (0, 1), (-1, -1), 1.6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 1.6),
            ('LEFTPADDING', (0, 0), (-1, -1), 3.5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3.5),
            ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.HexColor("#1E3A8A")),
            ('LINEBELOW', (0, 1), (-1, -1), 0.4, colors.HexColor("#F1F5F9")),
        ]

        for r_i in range(1, len(table_data)):
            if r_i % 2 == 0:
                t_style.append(('BACKGROUND', (0, r_i), (-1, r_i), colors.HexColor("#F8FAFC")))
            else:
                t_style.append(('BACKGROUND', (0, r_i), (-1, r_i), colors.white))

        t_itens.setStyle(TableStyle(t_style))
        
        story.append(t_sec)
        story.append(Spacer(1, 1.2 * mm))
        story.append(t_itens)

        if forcar_quebra_secao and idx < len(secoes_lista) - 1:
            story.append(PageBreak())
        else:
            story.append(Spacer(1, 4.5 * mm))

    doc.build(story, canvasmaker=NumberedCanvas)
    print(f"\nPDF gerado com sucesso: {caminho_pdf_saida} ({total_itens} itens incluídos)")
    return True


def main():
    parser = argparse.ArgumentParser(description="Gerador Inteligente de Relatório de Alteração de Preços")
    parser.add_argument("planilha", nargs="?", default="ALTERACAO_PRECO.xlsx", help="Caminho do arquivo XLSX")
    parser.add_argument("saida", nargs="?", default=None, help="Caminho do arquivo PDF de saída")
    parser.add_argument("--modo", choices=["primeiro", "segundo", "todos", "apos", "ate", "reset"], default="todos", 
                        help="Modo de geração: 'primeiro' (1º relatório), 'segundo' (2º relatório com apenas os novos), 'todos' (consolidado) ou 'reset'")
    parser.add_argument("--corte", default=None, help="Horário de corte manual (ex: 13:00)")
    parser.add_argument("--quebra-pagina-por-secao", action="store_true", help="Força cada seção a iniciar no topo de uma nova página")
    parser.add_argument("--forcar", action="store_true", help="Ignora a confirmação interativa de avisos")
    
    args = parser.parse_args()

    if args.modo == "reset":
        resetar_historico()
        return

    # 1. Checagem de Conformidade e Alertas
    alertas = verificar_conformidade_planilha(args.planilha)
    if alertas and not args.forcar:
        print("\n========================================================================")
        print("                  ⚠️   AVISO DE CONFORMIDADE DA PLANILHA   ⚠️")
        print("========================================================================")
        for a in alertas:
            print(f"\n{a}")
        print("========================================================================")
        print("  O padrão ideal é: Formato ANALÍTICO e SEM a seção de Hortifrúti.")
        print("========================================================================")
        try:
            resp = input("\nDeseja continuar mesmo assim e gerar o PDF? [S/N] (Enter = Sim): ").strip().upper()
            if resp == 'N':
                print("\nOperação cancelada pelo usuário.")
                print("Por favor, exporte no Varejofácil no formato ANALÍTICO e sem Hortifrúti.\n")
                sys.exit(0)
        except (EOFError, KeyboardInterrupt):
            pass

    dados = extrair_dados_xlsx(args.planilha)
    
    info_turno = ""
    if args.modo == "primeiro":
        info_turno = " (1º Relatório do Dia)"
        if args.saida is None:
            args.saida = "1_RELATORIO_PRECOS.pdf"
    elif args.modo == "segundo":
        info_turno = " (2º Relatório - Apenas Novas Alterações)"
        if args.saida is None:
            args.saida = "2_RELATORIO_PRECOS_NOVOS.pdf"
    else:
        if args.saida is None:
            args.saida = "RELATORIO_ALTERACAO_PRECO.pdf"

    secoes_filtradas, itens_filtrados = filtrar_dados(
        dados,
        modo=args.modo,
        hora_corte=args.corte
    )

    sucesso = gerar_pdf(
        dados,
        args.saida,
        secoes_filtradas,
        info_turno=info_turno,
        forcar_quebra_secao=args.quebra_pagina_por_secao
    )

    if sucesso and args.modo == "primeiro":
        data_hoje = dados['periodo'] if dados['periodo'] else datetime.now().strftime("%d/%m/%Y")
        if " a " in data_hoje:
            data_hoje = data_hoje.split(" a ")[-1].strip()
        salvar_historico(data_hoje, itens_filtrados, info_execucao=args.modo)


if __name__ == "__main__":
    main()
